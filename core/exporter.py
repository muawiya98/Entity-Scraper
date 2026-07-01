"""JSON export helpers.

Every completed search is written to ``data/json/`` as a single, human-readable
JSON file containing the search metadata and all scraped entities.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from config import config
from core import database


def _safe_slug(text: str, max_len: int = 40) -> str:
    slug = re.sub(r"[^\w؀-ۿ]+", "-", text, flags=re.UNICODE).strip("-")
    return (slug or "search")[:max_len]


def build_payload(search_id: int) -> dict:
    search = database.get_search(search_id) or {}
    entities = database.get_entities_for_search(search_id)
    return {
        "search": {
            "id": search.get("id"),
            "query": search.get("query"),
            "location": search.get("location"),
            "entity_type": search.get("entity_type"),
            "created_at": search.get("created_at"),
            "completed_at": search.get("completed_at"),
            "results_count": len(entities),
        },
        "entities": entities,
    }


def export_search(search_id: int) -> str:
    """Write the search payload to disk and return the file path."""
    payload = build_payload(search_id)
    query = payload["search"].get("query") or "search"
    filename = f"search_{search_id :04d}_{_safe_slug (query )}.json"
    path = Path(config.JSON_DIR) / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(path)


def export_excel(search_id: int) -> str:
    import openpyxl

    payload = build_payload(search_id)
    search = payload.get("search", {})
    entities = payload.get("entities", [])

    wb = openpyxl.Workbook()

    ws_entities = wb.active
    ws_entities.title = "Entities"
    ws_entities.append(
        [
            "Name",
            "Website",
            "Domain",
            "Description",
            "Address",
            "City",
            "Country",
            "Emails",
            "Phones",
            "Social Profiles",
        ]
    )

    ws_people = wb.create_sheet(title="People")
    ws_people.append(
        [
            "Entity Name",
            "Person Name",
            "Position",
            "Email",
            "Phone",
            "Profile URL",
            "Source",
        ]
    )

    ws_pages = wb.create_sheet(title="Pages")
    ws_pages.append(["Entity Name", "Page URL", "Page Type", "Status Code"])

    for ent in entities:

        emails = " | ".join(ent.get("emails") or [])
        phones = " | ".join(ent.get("phones") or [])
        social = " | ".join(f"{k }: {v }" for k, v in (ent.get("social") or {}).items())

        ws_entities.append(
            [
                ent.get("name") or "",
                ent.get("website") or "",
                ent.get("domain") or "",
                ent.get("description") or "",
                ent.get("address") or "",
                ent.get("city") or "",
                ent.get("country") or "",
                emails,
                phones,
                social,
            ]
        )

        for p in ent.get("people") or []:
            ws_people.append(
                [
                    ent.get("name") or "",
                    p.get("name") or "",
                    p.get("position") or "",
                    p.get("email") or "",
                    p.get("phone") or "",
                    p.get("profile_url") or "",
                    p.get("source") or "",
                ]
            )

        for page in ent.get("pages") or []:
            ws_pages.append(
                [
                    ent.get("name") or "",
                    page.get("url") or "",
                    page.get("page_type") or "",
                    page.get("status_code") or "",
                ]
            )

    query = search.get("query") or "search"
    filename = f"search_{search_id :04d}_{_safe_slug (query )}.xlsx"
    path = Path(config.JSON_DIR) / filename
    wb.save(path)

    return str(path)
